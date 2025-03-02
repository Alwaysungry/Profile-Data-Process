#按行读取文件1：
from itertools import islice
import pandas as pd;
import openpyxl as op
import numpy as np
from matplotlib import pyplot as plt
date1 = []
date2 = []
depth1 =[]
depth2 =[]
temperature1=[]
temperature2=[]
conductivity =[]
conductivity2 = []
f = open(r"H:\2021.8.25-8.28深蓝一号海试\204485_20210830_104214hour\204485_20210830_104214hour_data.txt","r")   #设置文件对象
for line in islice(f, 1, None):  # 去除txt文件的第一行
    temp = line.split(',')
    date1.append(temp[0])
    conductivity.append(temp[1])
    temperature1.append(temp[2])
    depth1.append(temp[5])
for i in range(len(conductivity)):
    if i%9000==0:
        depth2.append(depth1[i])
        temperature2.append(temperature1[i])
        conductivity2.append(conductivity[i])
        date2.append(date1[i])
print(temperature2)
# dfT = pd.DataFrame(temperature2, columns="temperature")
# dfC = pd.DataFrame(conductivity2, columns="conductivity")
#
# dfT.to_excel("H:\\lzy\\10minIntervaldata.xlsx",index=False)
# dfC.to_excel(r"H:\lzy\10minIntervaldata.xlsx",index=False)
bg = op.load_workbook(r"H:\\lzy\\10minIntervaldata.xlsx")      	# 应先将excel文件放入到工作目录下
sheet = bg["Sheet1"]                          		 	# “Sheet1”表示将数据写入到excel文件的sheet1下
for i in range(1, len(temperature2)+1):
    sheet.cell(i , 1, temperature2[i - 1])					# sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    sheet.cell(i, 2 ,conductivity2[i-1])
    sheet.cell(i, 3 ,depth2[i-1])
    sheet.cell(i,4,date2[i-1])
bg.save(r"H:\\lzy\\10minIntervaldata.xlsx")

# f = open(r"C:\Users\Hantong\Desktop\新建文本文档 (3).txt","r")   #设置文件对象
# for line in islice(f, 1, None):  # 去除txt文件的第一行
#     temp = line.split(',')
#     temperature2.append(temp[2])
#     depth2.append(temp[5])
# depth_array1 = np.array(depth1)
# depth_array1 = depth_array1.astype(np.float).tolist()
# depth_array2 = np.array(depth2)
# depth_array2 = depth_array2.astype(np.float).tolist()
# temperature_array_1 = np.array(temperature1)
# temperature_array_1 = temperature_array_1.astype(np.float).tolist()
# temperature_array_2 = np.array(temperature2)
# temperature_array_2 = temperature_array_2.astype(np.float).tolist()
#
# plt.plot(temperature_array_1, depth_array1)
#
# plt.plot(temperature_array_2, depth_array2)
# plt.ylabel('Depth(m)')
# plt.xlabel('Temperature(℃)')
# plt.gca().invert_yaxis()
# plt.show()